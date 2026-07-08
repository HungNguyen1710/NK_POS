from pathlib import Path
from openpyxl import load_workbook


class ExcelService:

    COLUMN_ALIASES = {
        "code": [
            "mã", "mã hàng", "mã sp", "mã sản phẩm",
            "sku", "code"
        ],

        "name": [
            "tên", "tên hàng", "tên sản phẩm",
            "product", "item"
        ],

        "unit": [
            "đvt",
            "đơn vị",
            "unit"
        ],

        "price": [
            "giá",
            "giá bán",
            "đơn giá",
            "selling price",
            "price"
        ],

        "stock": [
            "tồn",
            "tồn kho",
            "stock"
        ]
    }

    def __init__(self, excel_folder="excel"):
        self.excel_folder = Path(excel_folder)

    # ====================================================
    # File
    # ====================================================

    def get_excel_files(self):

        if not self.excel_folder.exists():
            return []

        return sorted(
            file
            for file in self.excel_folder.iterdir()
            if file.suffix.lower() in (".xlsx", ".xlsm")
        )

    # ====================================================
    # Workbook
    # ====================================================

    def load_workbook(self):

        files = self.get_excel_files()

        if not files:
            raise FileNotFoundError(
                "Không tìm thấy file Excel."
            )

        return load_workbook(
            files[0],
            data_only=True
        )

    # ====================================================
    # Product Sheet
    # ====================================================

    def find_product_sheet(self):

        workbook = self.load_workbook()

        best_sheet = None
        best_score = -1

        keywords = [
            "mã",
            "tên",
            "đvt",
            "giá",
            "tồn"
        ]

        for sheet in workbook.worksheets:

            score = 0

            for row in sheet.iter_rows(
                    min_row=1,
                    max_row=10):

                text = ""

                for cell in row:

                    if cell.value:

                        text += str(cell.value).lower()

                for keyword in keywords:

                    if keyword in text:
                        score += 1

            if score > best_score:
                best_score = score
                best_sheet = sheet

        return best_sheet

    # ====================================================
    # Header
    # ====================================================

    def find_header_row(self):

        sheet = self.find_product_sheet()

        best_row = 1
        best_score = -1

        for row in range(1, 21):

            text = ""

            for cell in sheet[row]:

                if cell.value:

                    text += str(cell.value).lower()

            score = 0

            for aliases in self.COLUMN_ALIASES.values():

                for alias in aliases:

                    if alias in text:
                        score += 1
                        break

            if score > best_score:
                best_score = score
                best_row = row

        return best_row

    # ====================================================
    # Mapping
    # ====================================================

    def get_column_mapping(self):

        sheet = self.find_product_sheet()

        header = self.find_header_row()

        mapping = {}

        for cell in sheet[header]:

            if cell.value is None:
                continue

            value = str(cell.value).lower().strip()

            for field, aliases in self.COLUMN_ALIASES.items():

                for alias in aliases:

                    if alias == value or alias in value:

                        mapping[field] = cell.column
                        break

        return mapping

    # ====================================================
    # Read Products
    # ====================================================

    def get_products(self):

        sheet = self.find_product_sheet()

        header = self.find_header_row()

        mapping = self.get_column_mapping()

        products = []

        row = header + 1

        while True:

            code_col = mapping.get("code")

            if code_col is None:
                break

            code = sheet.cell(row=row, column=code_col).value

            if code is None:
                break

            item = {
                "code": str(code).strip(),

                "name": self.get_value(
                    sheet,
                    row,
                    mapping.get("name")
                ),

                "unit": self.get_value(
                    sheet,
                    row,
                    mapping.get("unit")
                ),

                "price": self.to_number(
                    self.get_value(
                        sheet,
                        row,
                        mapping.get("price")
                    )
                ),

                "stock": self.to_number(
                    self.get_value(
                        sheet,
                        row,
                        mapping.get("stock")
                    )
                )
            }

            products.append(item)

            row += 1

        return products

    # ====================================================
    # Helpers
    # ====================================================

    def get_value(
            self,
            sheet,
            row,
            column):

        if column is None:
            return ""

        value = sheet.cell(
            row=row,
            column=column
        ).value

        if value is None:
            return ""

        return value

    def to_number(self, value):

        if value in ("", None):
            return 0

        try:
            return float(value)
        except:
            return 0

    # ====================================================
    # Debug
    # ====================================================

    def print_products(self):

        products = self.get_products()

        print()

        print("PRODUCTS")

        print("-----------------------------")

        for product in products:

            print(product)

        print()

        print("TOTAL :", len(products))